from pathlib import Path

import openpyxl

from utilities.excel_utils import read_configuration_data_from_excel


# need to define data for dev or qa
class HomePageData:
    message = "Success"
    name = "Abdullah Al Faroque"
    email = "aalfaroque@gmail.com"
    password = "123456"
    txt = "Hello Again"
    success_message = "Success"
    test_homepage_data = [
        {
            "first_name": "Abdullah Al",
            "email": "aalfaroque@gmail.com",
            "password": "12345",
        },
        {"first_name": "Faroque", "email": "rfnshare@gmail.com", "password": "4321"},
    ]
    EXCEL_PATH = Path(__file__).parent.parent / "test_data.xlsx"
    configuration_data = read_configuration_data_from_excel(EXCEL_PATH)
    url = configuration_data["frontend_url"]
    h_mode = configuration_data["headless"]
    browser = configuration_data["browser"].casefold()
    b = []
    if browser == "all":
        b = ["chrome", "firefox", "edge"]
    else:
        b.append(configuration_data["browser"].casefold())

    @staticmethod
    def get_test_data(TestCase):
        EXCEL_PATH = Path(__file__).parent / "PythonDemo.xlsx"
        book = openpyxl.load_workbook(EXCEL_PATH)
        sheet = book.active
        dic = {}
        keys = []
        values = [[] for i in range(1, sheet.max_column)]
        k = 0
        for i in range(1, 2):
            for j in range(1, sheet.max_column + 1):
                keys.append(sheet.cell(row=i, column=j).value)
        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == TestCase:
                for j in range(1, sheet.max_column + 1):
                    values[k].append(sheet.cell(row=i, column=j).value)
            k += 1
        lst_dic = []
        for i in values:
            dic = dict(zip(keys, i))
            if dic:
                lst_dic.append(dic)
        return lst_dic
